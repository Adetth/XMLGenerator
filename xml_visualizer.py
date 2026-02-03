from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element
from  dataclasses import dataclass
from enum import Enum
from typing import List

#I believe the way that the xml is made is that, whenever a color or some parameter is set,
#it creates a id for that color / font / border and stores it in the values section (for color and font)
#and the obj section (for borders), the numbering starts from the 16 bit max (2^15) being 32768
# so saying i want to add a color, i have to potentially increment this number by how many ever such colors
#already exist and i have to add the color there

#----------------DATA DECLARATION START----------------

form_xml_data = {
    "directory" : str, #format -> Parent/Child
    "name" : str
}

class DimType(Enum):
    MEMBER = "Member"
    FORMULA = "Formula"

#Making this a class since need many instances of this validation rule, due to many rules existing
#Not found better method yet...
@dataclass
class DimContainer:
    """
    Unified container for both Rows and Columns.
    """
    name: str
    size: str
    formula_or_function: DimType
    
    # Specifics
    formula_label: str 
    formula_equation: str
    function_name: str
    function_member_name: str   

#Dont care about POV, Business Rules, Display Options

@dataclass
class ValidationRuleContainer:
    
    #Validation rules are independent of the dimensions, they are assigned only to the grid
    #The style ID acts as a key between all the different tags    
    
    #Validation Rule
    name : str
    row_location : str #-1 is metadata, 0, 1, 2... are the rows themselves
    col_location :str
    
    #Validation Condition (inside Validation Rule)
    style_id : str
    bg_color : int #In decimal format, should be converted to hex for excel

@dataclass
class FormFormattingContainer:
    
    #this section basically assigns the formatting to the member / dim,
    #instead of it being restricted to row loc and col loc as it is laid out in the Validation Rules
    #this enables us to move around rows while the formatting moves with the row
    
    #I think i just have to ensure the segments and the style id remain the same,
    #since not able to see any difference
    
    style_id : str
    row_name : str
    row_segment : str
    row_dim : str
    col_name : str
    col_segment : str
    col_dim : str

@dataclass
class CellStyleContainer:
    
    #dictate how the cell's font, color, borders should look
    #again, linked by Style ID
    style_id : str
    font_id : str
    background_color_id : str
    
    #format tag may not exist for all, in that case make it nothing
    format_id : str
    
    #search for the obj tag inside the objs, if found, then add it to the list for borders
    #else move next
    border_id : list[str]

@dataclass
class ColorValueContainer:
    
    color_id : str
    
    red : str
    green : str
    blue : str

#not worrying about font for now

@dataclass
class TextFormatContainer:
    
    format_id : str
    value : str # bold italic or whatever

@dataclass
class BorderValueContainer:
    
    border_id : str
    red : str
    green : str
    blue : str
    placement : str
    style : str
    width : str

#----------------DATA DECLARATION END----------------

#-------------------DATA GET START-------------------

#2 types of data I need to get, one is the id inside the tag, and the other is the content between

#arrow means nothing to python, only for the me and vscode to understand
#called a "Return Type Hint" to show us what the return type should look like
#making a list of instanciated dataclass elements
def get_dim_containers(root, xml_path: str) -> List[DimContainer]:
    """
    Parses either Rows or Columns based on the xml_path provided.
    """
    container_list = []
    
    # Dynamic path: find either <rows> or <columns>
    parent_node = root.find(xml_path)
    
    if parent_node is None:
        return []
    
    for segment in parent_node.findall("segment"):
        # Common attribute: Size (Height for Rows, Width for Cols)
        # Note: Columns have width="...", Rows have size="..." or height="..."
        # For simplicity in this unified class, we just grab 'size' or 'width'
        dim_size = segment.get("size", segment.get("width", ""))
        
        dimension = segment.find("dimension")
        if dimension is None: continue

        # Init variables
        d_name = ""
        d_type = None
        d_formula_label = ""
        d_equation = ""
        d_func_name = ""
        d_func_mbr = ""
        
        formula_data = dimension.find("formula")
        function_data = dimension.find("function")
        
        if formula_data is not None:
            d_type = DimType.FORMULA
            d_formula_label = formula_data.get("label", "")
            d_equation = formula_data.get("formulaValue", "")
            d_name = d_formula_label 
        
        elif function_data is not None:
            d_type = DimType.MEMBER
            d_func_name = function_data.get("name", "")
            
            member_node = function_data.find("member")
            if member_node is not None:
                d_func_mbr = member_node.get("name", "")
                d_name = d_func_mbr 
        
        if d_type:
            new_dim = DimContainer(
                name=d_name,
                size=dim_size,
                formula_or_function=d_type,
                formula_label=d_formula_label,
                formula_equation=d_equation,
                function_name=d_func_name,
                function_member_name=d_func_mbr
            )
            container_list.append(new_dim)

    return container_list
def get_validation_containers(root) -> List[ValidationRuleContainer]:
    val_list = []
    val_node = root.find(".//dataValidationRules")
    
    if val_node is None:
        return []

    for val_rule in val_node.findall("dataValidationRule"):
        
        val_name = val_rule.get("name")
        row_loc = val_rule.get("rowLocation")
        col_loc = val_rule.get("colLocation")
        
        val_cond = val_rule.find("dataValidationCond")
        if val_cond is None: 
            continue
        
        val_style_id = val_cond.get("styleId")
        val_bg_decimal = val_cond.get("bgColor") 
        
        final_bg_color = None
        if val_bg_decimal:
            try:
                # FIX: Force 6 digits with leading zeros (e.g. 10 -> "00000A")
                # :06X means "6 digits, Hexadecimal, Uppercase"
                final_bg_color = f"{int(val_bg_decimal):06X}"
            except ValueError:
                pass 

        if val_name:
            new_rule = ValidationRuleContainer(
                name = val_name,
                row_location = row_loc,
                col_location = col_loc,
                style_id = val_style_id,
                bg_color = final_bg_color
            )
            val_list.append(new_rule)
    
    return val_list
def get_form_formatting_container(root) -> List[FormFormattingContainer]:
    
    form_list = []
    tuples_node = root.find(".//formFormattings/formFormatting/dataCellMbrTuples")
    
    if tuples_node is None:
        return []
    
    for rule in tuples_node.findall("dataCellMbrTuple"):
        
        style_node = rule.find("cellStyleId")
        if style_node is None: 
            continue
        current_style_id = style_node.text
        
        r_name = r_seg = r_dim = ""
        c_name = c_seg = c_dim = ""
        
        for mbr_tuple in rule.findall("frmMbrTuple"):
            
            loc_node = mbr_tuple.find("gridLocation")
            member_node = mbr_tuple.find("mbr")
            
            if loc_node is None or member_node is None: 
                continue

            location_text = loc_node.text
            
            if location_text == "rows":
                r_name = member_node.get("name")
                r_seg = member_node.get("segment")
                r_dim = member_node.get("dim")
                
            elif location_text == "columns":
                c_name = member_node.get("name")
                c_seg = member_node.get("segment")
                c_dim = member_node.get("dim")
        
        if r_name and c_name:
            new_fmt = FormFormattingContainer(
                style_id = current_style_id,
                row_name = r_name,
                row_segment = r_seg,
                row_dim = r_dim,
                col_name = c_name,
                col_segment = c_seg,
                col_dim = c_dim
            )
            form_list.append(new_fmt)
    
    return form_list
def get_color_containers(root) -> List[ColorValueContainer]:
    
    color_list = []
    # Locate the colors section inside values
    colors_node = root.find(".//values/colors")
    
    if colors_node is None:
        return []
    
    for color in colors_node.findall("color"):
        
        # Get attributes directly from the tag
        c_id = color.get("id")
        r = color.get("R")
        g = color.get("G")
        b = color.get("B")
        
        if c_id:
            new_color = ColorValueContainer(
                color_id = c_id,
                red = r,
                green = g,
                blue = b
            )
            color_list.append(new_color)
            
    return color_list
def get_text_format_containers(root) -> List[TextFormatContainer]:
    
    fmt_list = []
    # Locate text formats section
    txt_node = root.find(".//values/txtFrmts")
    
    if txt_node is None:
        return []

    for fmt in txt_node.findall("txtFrmt"):
        
        # ID is an attribute
        f_id = fmt.get("id")
        
        # Value is the text content between tags
        # Use strip() to remove accidental whitespace
        f_val = fmt.text.strip() if fmt.text else ""
        
        if f_id:
            new_fmt = TextFormatContainer(
                format_id = f_id,
                value = f_val
            )
            fmt_list.append(new_fmt)
            
    return fmt_list
def get_border_containers(root) -> List[BorderValueContainer]:
    
    border_list = []
    # Locate borders section inside objs
    borders_node = root.find(".//objs/borders")
    
    if borders_node is None:
        return []

    for border in borders_node.findall("border"):
        
        # Find child tags for properties
        id_node = border.find("id")
        place_node = border.find("placement")
        style_node = border.find("style")
        width_node = border.find("width")
        
        # Color is a nested tag with attributes inside the border tag
        color_node = border.find("color")
        
        # Initialize color vars
        r = g = b = "0"
        if color_node is not None:
            r = color_node.get("R", "0")
            g = color_node.get("G", "0")
            b = color_node.get("B", "0")
            
        # Get text content safely from the child nodes
        b_id = id_node.text if id_node is not None else ""
        place = place_node.text if place_node is not None else ""
        style = style_node.text if style_node is not None else ""
        width = width_node.text if width_node is not None else ""

        if b_id:
            new_border = BorderValueContainer(
                border_id = b_id,
                red = r,
                green = g,
                blue = b,
                placement = place,
                style = style,
                width = width
            )
            border_list.append(new_border)
            
    return border_list
def get_cell_style_containers(root) -> List[CellStyleContainer]:
    
    style_list = []
    # Locate cell styles section
    styles_node = root.find(".//cellStyles")
    
    if styles_node is None:
        return []

    for style in styles_node.findall("cellStyle"):
        
        s_id = style.get("id")
        
        # Locate the values sub-section
        vals = style.find("cellStyleValues")
        if vals is None: continue
        
        # Extract component IDs (tags with 'id' attributes)
        
        # Font
        font_node = vals.find("font")
        font_id = font_node.get("id") if font_node is not None else ""
        
        # Background Color
        bg_node = vals.find("backColor")
        bg_id = bg_node.get("id") if bg_node is not None else ""
        
        # Text Format (Optional - e.g. Bold)
        fmt_node = vals.find("format")
        fmt_id = fmt_node.get("id") if fmt_node is not None else ""
        
        # Borders: List of IDs
        # Look inside the <objs> tag of THIS style
        border_ids = []
        objs_node = style.find("objs")
        if objs_node is not None:
            # Oracle lists borders as <obj type="border" id="32768" />
            for obj in objs_node.findall("obj"):
                if obj.get("type") == "border":
                    border_ids.append(obj.get("id"))
        
        if s_id:
            new_style = CellStyleContainer(
                style_id = s_id,
                font_id = font_id,
                background_color_id = bg_id,
                format_id = fmt_id,
                border_id = border_ids
            )
            style_list.append(new_style)
            
    return style_list

#Getting the form data from the XML file
# INPUT_XML_FILE = "Adetth - Segment Properties Changed.xml"
# INPUT_XML_FILE = "Adetth - 1. Control Form.xml"
INPUT_XML_FILE = "MPB_1.2 Manage New Hires.xml"

tree = ET.parse(INPUT_XML_FILE)
root = tree.getroot()

def get_form_xml_data():
    form_xml_data["directory"] = root.get("dir")
    form_xml_data["name"] = root.get("name")
    
    return form_xml_data

#----------USER VARS TO CHANGE NAME ETC-----------

xml_file_location = "Adetth - Segment Properties Changed.xml"
output_file = f"XML_REPORT_FOR_{xml_file_location[:-4]}.xlsx"

def generate_excel_report(root, output_file):
    
    rows = get_dim_containers(root, ".//query/rows")
    cols = get_dim_containers(root, ".//query/columns")
    vals = get_validation_containers(root)
    frmt = get_form_formatting_container(root)
    color = get_color_containers(root)
    text = get_text_format_containers(root)
    border = get_border_containers(root)
    cstyle = get_cell_style_containers(root)
    
    wb = Workbook()
    
    #to make it so that i dont have the keep declaring the header format for the Excel file
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    def setup_sheet(ws, headers):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    #Sheet 1 : has the details of the rows and columns, checks if they are formula rows
    # Category : Row / Col
    # Name : Whats displayed in the text box
    # Type : Either a Forumla row/col or a Member
    # Size : size of the row/col
    # Formula / Member Details : Shows the formula / details of the formula row/col
    
    current_row = 2 #start from 2nd row since we have the header row already 
    
    setup_sheet(wb.active, ["Test 1", "TEST 0                    2"])
    wb.save(output_file)
    print(f"{output_file} Saved!")
    
    

# --- EXECUTE ---
if __name__ == "__main__":
    generate_excel_report(root, output_file)